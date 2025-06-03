import os
from openpyxl import load_workbook
from .table_utils import merge_and_deduplicate


def process_processed_files(directory):
    """
    对指定目录中所有xlsx文件的所有工作表执行：
    - 合并并去重多行表头
    - 将结果覆盖回原工作表
    保存修改。
    """
    for fname in os.listdir(directory):
        if not fname.lower().endswith('.xlsx'):
            continue
        fullpath = os.path.join(directory, fname)
        wb = load_workbook(fullpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            processed = merge_and_deduplicate(rows)
            # 写回工作表
            for row_idx, row in enumerate(processed):
                for col_idx, cell in enumerate(row):
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell)
        wb.save(fullpath)


def remove_empty_columns(directory):
    """
    删除指定目录所有xlsx文件中，所有工作表的空白列。
    空白列定义为整列数据均为空或空字符串。
    删除后新建工作表替换原工作表。
    """
    for fname in os.listdir(directory):
        if not fname.lower().endswith('.xlsx'):
            continue
        fullpath = os.path.join(directory, fname)
        wb = load_workbook(fullpath, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            num_cols = len(rows[0])
            non_empty_col_indices = []

            # 找出非空列索引
            for col_idx in range(num_cols):
                is_column_empty = all(
                    row[col_idx] is None or str(row[col_idx]).strip() == ''
                    for row in rows
                )
                if not is_column_empty:
                    non_empty_col_indices.append(col_idx)

            # 若全列均非空，无需修改
            if len(non_empty_col_indices) == num_cols:
                continue

            # 创建新工作表，写入非空列数据
            new_ws = wb.create_sheet(f"{sheet_name}_temp")
            for r_idx, row in enumerate(rows):
                new_row = [row[i] for i in non_empty_col_indices]
                new_ws.append(new_row)

            # 删除旧表，重命名新表
            del wb[sheet_name]
            new_ws.title = sheet_name

        wb.save(fullpath)


def deduplicate_column_names(directory):
    """
    对指定目录下所有xlsx文件的所有工作表的第一行列名去重，
    通过在重复列名后添加空格数量区分。
    保存修改回文件。
    """
    for fname in os.listdir(directory):
        if not fname.lower().endswith('.xlsx'):
            continue
        fullpath = os.path.join(directory, fname)
        wb = load_workbook(fullpath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = list(rows[0])
            seen = {}
            new_header = []
            for col in header:
                base = col if col is not None else ''
                if base not in seen:
                    seen[base] = 0
                    new_header.append(base)
                else:
                    seen[base] += 1
                    # 通过添加空格数量实现列名去重
                    new_header.append(f"{base}{' ' * seen[base]}")

            # 更新工作表第一行列名
            for col_idx, val in enumerate(new_header):
                ws.cell(row=1, column=col_idx + 1, value=val)

        wb.save(fullpath)
