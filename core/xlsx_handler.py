import os
from openpyxl import load_workbook
from .table_utils import extract_tables_from_rows, transform_table


def fill_merged_cells(sheet):
    """
    对 openpyxl 的工作表中所有合并单元格进行拆分，
    并将合并区域内所有单元格填充为合并区域左上角的值。
    """
    for merged in list(sheet.merged_cells.ranges):
        # 先取消合并
        sheet.unmerge_cells(range_string=str(merged))
        min_col, min_row, max_col, max_row = merged.bounds
        # 取左上角单元格的值
        value = sheet.cell(row=min_row, column=min_col).value
        # 填充合并区域内所有单元格
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                sheet.cell(row=r, column=c).value = value


def process_xlsx(path):
    """
    处理xlsx文件，针对每个工作表：
    - 填充合并单元格
    - 读取所有行
    - 提取多个表格块
    - 对每个表格块执行表头转换
    返回所有表格及其所属工作表名。
    """
    wb = load_workbook(path, data_only=True)
    all_tables = []
    for sheet in wb.worksheets:
        fill_merged_cells(sheet)
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        tables = extract_tables_from_rows(rows)
        for tbl in tables:
            processed = transform_table(tbl)
            all_tables.append((sheet.title, processed))
    return all_tables
